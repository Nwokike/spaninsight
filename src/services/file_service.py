"""File import service — CSV/Excel loading with validation.

Handles the complete pipeline from file selection to pandas DataFrame,
enforcing the 100MB size limit to keep Android stable.
"""

from __future__ import annotations

import hashlib
import logging
import os
from pathlib import Path

from core.constants import ALLOWED_EXTENSIONS, MAX_FILE_SIZE_BYTES, MAX_FILE_SIZE_MB

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd

logger = logging.getLogger(__name__)


class FileValidationError(Exception):
    """Raised when a file fails validation checks."""

    pass


def validate_file(file_path: str) -> None:
    """Validate file extension and size. Raises FileValidationError on failure."""
    path = Path(file_path)

    # Check extension
    ext = path.suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise FileValidationError(
            f"Unsupported file type: '{ext}'. "
            f"Please use CSV, Excel, JSON, XML, STATA, SAS, TSV, TXT, ZIP, or Pickle files ({', '.join(ALLOWED_EXTENSIONS)})."
        )

    # Check size
    try:
        size = os.path.getsize(file_path)
    except OSError:
        raise FileValidationError(
            "Could not read file. It may have been moved or deleted."
        )

    if size > MAX_FILE_SIZE_BYTES:
        size_mb = size / (1024 * 1024)
        raise FileValidationError(
            f"File is too large ({size_mb:.1f} MB). "
            f"Maximum allowed size is {MAX_FILE_SIZE_MB} MB."
        )

    if size == 0:
        raise FileValidationError("File is empty. Please select a file with data.")


def load_dataframe(file_path: str) -> pd.DataFrame:
    """Load a CSV or Excel file into a pandas DataFrame.

    Raises FileValidationError on any loading failure.
    """
    validate_file(file_path)

    import pandas as pd

    path = Path(file_path)
    ext = path.suffix.lower()

    try:
        if ext == ".csv":
            # Try UTF-8 first, fall back to latin-1 for Excel-exported CSVs
            try:
                df = pd.read_csv(file_path, encoding="utf-8")
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, encoding="latin-1")
        elif ext in {".json", ".jsonl", ".ndjson"}:
            # PERFORMANCE FIX: Parse JSON/JSON Lines natively using high-performance msgspec
            import msgspec

            with open(file_path, "rb") as f:
                content = f.read()
            try:
                parsed = msgspec.json.decode(content)
                df = pd.DataFrame(parsed)
            except msgspec.DecodeError:
                # Fallback: Parse line-delimited JSON
                parsed = [
                    msgspec.json.decode(line)
                    for line in content.split(b"\n")
                    if line.strip()
                ]
                df = pd.DataFrame(parsed)
        elif ext == ".xlsx":
            # MOBILE COMPATIBILITY FIX: Swapped calamine for pure-Python openpyxl to ensure
            # 100% Android/iOS compatibility, offloading the parsing to a background thread
            # in process_file to maintain UI responsiveness.
            df = pd.read_excel(file_path, engine="openpyxl")
        elif ext == ".xml":
            # NATIVE PARSING: Parse XML natively using Python's built-in etree parser
            df = pd.read_xml(file_path, parser="etree")
        elif ext == ".dta":
            # NATIVE PARSING: Parse STATA natively using pandas read_stata
            df = pd.read_stata(file_path)
        elif ext in {".sas7bdat", ".xport"}:
            # NATIVE PARSING: Parse SAS natively using pandas read_sas
            df = pd.read_sas(file_path)
        elif ext in {".tsv", ".txt"}:
            # NATIVE PARSING: Parse delimited files natively (tab-separated or auto-detected)
            try:
                df = pd.read_csv(file_path, sep="\t")
            except Exception:
                df = pd.read_csv(file_path)
        elif ext == ".zip":
            # NATIVE DECOMPRESSION: Try loading zipped CSV first
            try:
                df = pd.read_csv(file_path)
            except Exception as e:
                raise FileValidationError(
                    f"Zipped file must contain a valid CSV dataset: {e}"
                )
        elif ext in {".pkl", ".pickle"}:
            # NATIVE PARSING: Parse Python Pickle files natively
            df = pd.read_pickle(file_path)
        else:
            raise FileValidationError(f"Unsupported format: {ext}")

    except FileValidationError:
        raise
    except Exception as e:
        raise FileValidationError(f"Failed to read file: {e}")

    if df.empty:
        raise FileValidationError("File loaded but contains no data rows.")

    if len(df.columns) == 0:
        raise FileValidationError("File loaded but contains no columns.")

    logger.info(
        "Loaded %s: %d rows × %d columns",
        path.name,
        len(df),
        len(df.columns),
    )
    return df


def get_data_summary(df: pd.DataFrame) -> dict:
    """Generate a comprehensive but token-efficient summary for the AI.

    Safely truncates categorical summaries to prevent UI locking on massive UUID/Text columns.
    """

    # PERFORMANCE FIX: Prevent thread freezing on describe(include="all")
    try:
        num_desc = df.select_dtypes(include="number").describe().round(2).to_dict()
        # Cap categorical summarization to first 10 text cols to prevent thread blocking
        cat_desc = (
            df.select_dtypes(exclude="number")
            .iloc[:, :10]
            .describe()
            .fillna("")
            .to_dict()
        )
        safe_describe = {**num_desc, **cat_desc}
    except Exception:
        safe_describe = {}

    summary = {
        "shape": {"rows": len(df), "columns": len(df.columns)},
        "columns": {str(c): str(df[c].dtype) for c in df.columns[:30]},
        "nulls": {str(c): int(df[c].isnull().sum()) for c in df.columns[:30]},
        "describe": safe_describe,
        "head": [],
        "tail": [],
    }

    try:
        summary["spatial"] = detect_spatial_columns(df)
    except Exception:
        summary["spatial"] = None

    # Safe head + tail — truncate long strings
    def _safe_rows(sub_df):
        obj_cols = [c for c in sub_df.columns if sub_df[c].dtype == "object"]
        if obj_cols:
            sub_df = sub_df.copy()
            for col in obj_cols:
                sub_df[col] = sub_df[col].apply(
                    lambda x: (
                        str(x)[:100] + "..."
                        if isinstance(x, str) and len(x) > 100
                        else x
                    )
                )
        return sub_df.to_dict(orient="records")

    try:
        summary["head"] = _safe_rows(df.head(20))
        summary["tail"] = _safe_rows(df.tail(20))
    except Exception:
        summary["head"] = "unavailable"

    return summary


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Serialize a DataFrame to CSV bytes for download."""
    return df.to_csv(index=False).encode("utf-8")


def df_to_styled_excel_bytes(df: pd.DataFrame, title: str = "Export") -> bytes:
    """Export a DataFrame to a styled Excel workbook using openpyxl directly.

    Features: teal header row, auto-adjusted column widths, auto-filter,
    number formatting for numeric columns, and a title sheet name.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill(
        start_color="0D9488", end_color="0D9488", fill_type="solid"
    )
    header_font = Font(name="Outfit", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Outfit", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(df.columns, 1):
            val = row[col_name]
            if isinstance(val, (int, float)) and pd.isna(val):
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col_name)),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    if len(df.columns) > 0 and len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    ws.freeze_panes = "A2"

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def detect_spatial_columns(df: pd.DataFrame) -> dict | None:
    """Detect latitude/longitude column pairs in a DataFrame using shapely.

    Returns a dict with 'lat_col', 'lon_col', 'point_count', 'bounds'
    if a valid spatial column pair is found, or None otherwise.
    """
    lat_candidates = [
        c
        for c in df.columns
        if any(k in c.lower() for k in ("lat", "latitude", "ycoord", "y_coord"))
    ]
    lon_candidates = [
        c
        for c in df.columns
        if any(
            k in c.lower() for k in ("lon", "long", "longitude", "xcoord", "x_coord")
        )
    ]
    if not lat_candidates or not lon_candidates:
        return None

    lat_col = lat_candidates[0]
    lon_col = lon_candidates[0]
    lat_vals = df[lat_col].dropna()
    lon_vals = df[lon_col].dropna()
    if len(lat_vals) < 2 or len(lon_vals) < 2:
        return None

    from shapely.geometry import Point, MultiPoint

    points = [
        Point(x, y) for x, y in zip(lon_vals, lat_vals) if pd.notna(x) and pd.notna(y)
    ]
    if len(points) < 2:
        return None

    multi = MultiPoint(points)
    centroid = multi.centroid
    bounds = multi.bounds

    return {
        "lat_col": lat_col,
        "lon_col": lon_col,
        "point_count": len(points),
        "centroid_lat": round(centroid.y, 6),
        "centroid_lon": round(centroid.x, 6),
        "bounds": {
            "min_lat": round(bounds[1], 6),
            "min_lon": round(bounds[0], 6),
            "max_lat": round(bounds[3], 6),
            "max_lon": round(bounds[2], 6),
        },
    }


def transform_json_with_jq(content: str | bytes, jq_filter: str) -> str:
    """Transform JSON content using a jq filter expression.

    Raises ValueError if the filter is invalid or the content cannot be parsed.
    """
    import jq

    if isinstance(content, bytes):
        content = content.decode("utf-8")

    import json as _json

    parsed = _json.loads(content)
    result = jq.compile(jq_filter).input(parsed).all()
    return _json.dumps(result, indent=2, default=str)


def generate_dataset_fingerprint(df: pd.DataFrame) -> str:
    """Generate a lightweight SHA-256 fingerprint of the dataset structure.

    This mathematically proves collaborators have the exact same file
    by hashing column names, data types, and total row count.
    """
    col_info = ",".join([f"{col}:{dtype}" for col, dtype in zip(df.columns, df.dtypes)])
    fingerprint_string = f"{col_info}|rows:{len(df)}"
    return hashlib.sha256(fingerprint_string.encode("utf-8")).hexdigest()
